import time
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from openpyxl import load_workbook

#Load the excelsheet
workbook = load_workbook("file.xlxs")

sheet = workbook.active
driver= webdriver.Chrome(ChromeDriverManager().install())
driver.maximize_window()
for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row, values_only=True):
    username= row[0]
    password= row[1]
    driver.ger("https://www.saucedemo.com/")

    time.sleep(3)
    driver.find_element(By.ID,"user-name").send_keys(username)
    driver.find_element(By.ID,"password").send_keys(password)
    driver.find_element(By.ID,"login-button").click()
    time.sleep(3)
driver.close()
